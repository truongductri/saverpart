# -*- coding: utf-8 -*-
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.worksheet.datavalidation import DataValidation
#from openpyxl.workbook import defined_name
import datetime
import quicky
from format_worksheet import worksheet_style, format_style 
import constant as KEY 
from performance.lv_core import db
import manager


@quicky.view.template("")
def call(request):
    token_key = request._get_get().get("tk")
    token_data = manager.get_parameter(request, token_key)
    
    if token_data.get('data', None) is None:
        return HttpResponse(content= "<h3>" + token_data["error"] + "</h3>", mimetype='text/html')

    collection_name = token_data["data"]["collection_name"]
    wb = Workbook()
    ws = wb.active
    #ws["A1"].value = "TEST"
    ws.title = KEY.WS_PREFIX_NAME + "1";

    #add sheet mapping
    wb.create_sheet(KEY.WS_MAPPING_SHEET);
    ws_mapping = wb[KEY.WS_MAPPING_SHEET];

    #add sheet init data
    wb.create_sheet(KEY.WS_INIT_DATA);
    ws_init = wb[KEY.WS_INIT_DATA];

    #Get column config 
    ret_columns = db.get_collection("HCSSYS_CollectionInfo").aggregate([
            {
                '$match': { 
                    '$and' : [
                        {'parent_field' : None}, 
                        {'field_path': {'$regex': '^,' + collection_name + ','}},
                        {'is_parent' : False}, 
                    ] 
                }
            },
            {'$project': {'_id': 0}}
        ]);
    columns = list(ret_columns)

    #Get export data
    ret_data = db.get_collection(collection_name).find({})
    data_items = list(ret_data);

    if(len(data_items) == 0):
        raise "Collection '" + collection_name + "' has not been declared in the database"
        return null;

    ref_accessmode = [{
        'value':1,
        'caption': u'Toàn quyền'
    },{
        'value': 2,
        'caption': u'Từ vùng được chọn đến các cấp con'
    },{
        'value': 3,
        'caption': u'Chỉ trên vùng được chọn'
    }]

    #create reference sheet
    ref_columns = [a for a in columns if a["field_name"] == "access_mode"]
    for col in ref_columns:
        wb.create_sheet(KEY.WS_PREFIX_REFERENCE + col["field_name"])
        ws_ref = wb[KEY.WS_PREFIX_REFERENCE + col["field_name"]]

        for iData, vData in enumerate(ref_accessmode):
            ws_ref.cell(row=iData+1, column=1).value = vData["value"]
            ws_ref.cell(row=iData+1, column=2).value = vData["caption"]
            ws_ref.cell(row=iData+1, column=3).value = vData["value"]


    
    #create header title
    header_fields = []
    for c in columns:
        if(c["field_name"] == 'access_mode'):
            header_fields.append({
                'field_name': c["field_name"],
                'key' : KEY.PREFIX_INIT,
                'display_name': c["description"],
                'is_ref' : True,
                'is_data_field' : False,
                'is_hidden' : True
            })
            header_fields.append({
                'field_name': c["field_name"],
                'key' : KEY.PREFIX_REF,
                'display_name': c["description"],
                'is_ref' : True,#List of reference data
                'is_data_field' : False,
                'is_hidden' : False
            })
            header_fields.append({
                'field_name': c["field_name"],
                'key' : c["field_name"],
                'display_name': c["description"],
                'is_ref' : True,
                'is_data_field' : True, #data_field in database
                'is_hidden' : True
            })
        else:
            header_fields.append({
                'field_name': c["field_name"],
                'key' : c["field_name"],
                'display_name': c["description"],
                'is_ref' : False,
                'is_data_field' : True,
                'is_hidden' : False
            })

    #Create header worksheet (row 1)
    idx_mapping_row_column = 1
    ws_mapping.cell(row=1, column=1).value = KEY.BEGIN_MAPPING
    ws_mapping.cell(row=1, column=2).value = collection_name
    ws_mapping.cell(row=1, column=3).value = ws.title
    idx_mapping_row_column += 1

    idx_init_column = 0
    for  iCol, vCol in enumerate(header_fields):
        if not (vCol["key"] == KEY.PREFIX_INIT):
            cell = ws.cell(row=1, column=iCol + 1 - idx_init_column)
            cell.value = vCol["display_name"]

            ws.column_dimensions[cell.column].width = len(vCol["display_name"]) if len(vCol["display_name"]) > 20 else 20 #20 characters
            if(vCol["is_hidden"]):
                ws.column_dimensions[cell.column].hidden = True

            if(vCol["is_data_field"] == True):
                #create mapping data
                ws_mapping.cell(row=idx_mapping_row_column, column=1).value = cell.column
                ws_mapping.cell(row=idx_mapping_row_column, column=2).value = vCol["field_name"]
                idx_mapping_row_column += 1
        else:
            cell = ws_init.cell(row=1, column=idx_init_column + 1)
            cell.value = vCol["field_name"]
            idx_init_column += 1

    ws_mapping.cell(row=idx_mapping_row_column, column=1).value = KEY.END_MAPPING

    #Render content to worksheet
    #if (len(data_items) > 0):
    for iItem, vItem in enumerate(data_items):
        num_init_column = 0
        for iCol, vCol in enumerate(header_fields):
            idx_col = iCol - num_init_column
            curr_cell = ws.cell(row = iItem + 2, column = idx_col + 1)
            if (vCol["is_data_field"] and not vCol["is_ref"]) :
                curr_cell.value = vItem[vCol["field_name"]]
            elif (vCol["is_ref"] and vCol["key"] == KEY.PREFIX_INIT):
                init_cell = ws_init.cell(row = iItem + 2, column = num_init_column + 1)
                init_cell.value = vItem[vCol["field_name"]]
                num_init_column +=1
            elif (vCol["is_ref"] and vCol["key"] == KEY.PREFIX_REF):
                #curr_cell: value list
                #vlookup column 1-2 in reference data
                ws_ref = wb[KEY.WS_PREFIX_REFERENCE + col["field_name"]]
                #ref_beg_cell = ws_ref.cell(row=1, column=1)
                ref_end_cell = ws_ref.cell(row = ws_ref.max_row, column = 2)
                ref_address = KEY.WS_PREFIX_REFERENCE + col["field_name"] + "!" + \
                    '$A$1:$B$' + str(ws_ref.max_row)
                    # "$" + ref_beg_cell.column + "$" + str(ref_beg_cell.col_idx) + ":" + \
                    # "$" + ref_end_cell.column + "$" + str(ref_end_cell.col_idx)
                init_cell = ws_init.cell(row = iItem + 2, column = num_init_column)
                curr_cell.value = "=VLOOKUP(" + KEY.WS_INIT_DATA + "!" +  init_cell.coordinate + "," + ref_address+ ",2, FALSE)"

                ref_address_title = KEY.WS_PREFIX_REFERENCE + col["field_name"] + "!" + \
                    "$B$1:" + "$B$" + str(ws_ref.max_row)

                # Create a data-validation object with list validation
                dv = DataValidation(type="list", formula1="=" + ref_address_title, allow_blank=True)
                # Optionally set a custom error message
                dv.error ='Your entry is not in the list'
                dv.errorTitle = 'Invalid Entry'

                # Optionally set a custom prompt message
                dv.prompt = 'Please select from the list'
                dv.promptTitle = 'List Selection'
                ws.add_data_validation(dv)
                dv.add(curr_cell)
            elif  (vCol["is_data_field"] and vCol["is_ref"]):
                pre_cell = ws.cell(row = iItem + 2, column = idx_col)

                #vlookup column 2-3 in reference data
                ws_ref = wb[KEY.WS_PREFIX_REFERENCE + col["field_name"]]
                #ref_beg_cell = ws_ref.cell(row=1, column=2)
                ref_end_cell = ws_ref.cell(row = ws_ref.max_row, column = ws_ref.max_column)
                ref_address = KEY.WS_PREFIX_REFERENCE + col["field_name"] + "!" + \
                    '$B$1:$C$' + str(ws_ref.max_row)
                    #"$" + ref_beg_cell.column + "$" + str(ref_beg_cell.col_idx) + ":" + \
                    #"$" + ref_end_cell.column + "$" + str(ref_end_cell.col_idx)
                curr_cell.value = "=VLOOKUP(" + pre_cell.coordinate + "," + ref_address+ ",2, FALSE)"

    #format worksheet
    ws = format_style(worksheet_style["NORMAL"]).format(ws)

    #wb.defined_names.append(defined_name.DefinedName(attr_text="HCSSYS_DataDomain!A$1", name="TEST_NAME_0000000"))


    response = HttpResponse(content=save_virtual_workbook(wb), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    time_export = datetime.datetime.now().strftime("%Y.%m.%dT%H.%M.%S");
    response['Content-Disposition'] = 'attachment; filename=' + 'export-' + time_export + '.xlsx'
    return response
