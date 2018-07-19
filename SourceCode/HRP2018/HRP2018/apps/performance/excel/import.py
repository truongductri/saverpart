from django.http import HttpResponse
import quicky
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import load_workbook
import datetime
import base64
from io import BytesIO
import constant as KEY
from performance.lv_core import db

def call(args):
    wb = load_workbook(filename=BytesIO(base64.b64decode(args["data"]["_content"])), data_only=True)
    #check exists mapping sheet
    if KEY.WS_MAPPING_SHEET in wb.sheetnames:
        ws_mapping = wb[KEY.WS_MAPPING_SHEET]
        mapping_configs = __get_mapping(ws_mapping)
        data = []
        for map_item in mapping_configs:
            #ws_data = wb[map_item['sheet_name']]
            _data = __get_data(map_item, wb[map_item['sheet_name']])
            data.append(_data)

            db.Collection(_data["collection_name"]).insert_or_update(_data["documents"])

        return {'config': mapping_configs, 'data': _data}
    else:
        return {
            error: "File structure error!"
        }

def __get_data(mapping_config, ws_data):
    """Convert data worksheet to json"""
    ret_data = {
        'collection_name' : mapping_config['collection_name'],
        'documents': []
    }

    for idx_row in range(2, ws_data.max_row + 1):
        data_row = {};
        for config in mapping_config["fields"]:
            data_row[config['field_name']] = ws_data[config['column'] + str(idx_row)].value
        ret_data['documents'].append(data_row)
    return ret_data

def __get_mapping(ws_mapping):
    """Read data in mapping sheet and convert to python object"""
    mapping_configs = []
    config = {
        'collection_name' : None,
        'sheet_name' : None,
        'fields' : []
    }
    for idx_row in range(1, ws_mapping.max_row + 1):
        cell_col1 = ws_mapping.cell(row=idx_row, column=1)
        cell_col2 = ws_mapping.cell(row=idx_row, column=2)
        cell_col3 = ws_mapping.cell(row=idx_row, column=3)
        if(cell_col1.value == KEY.BEGIN_MAPPING):
            config['collection_name'] = cell_col2.value
            config['sheet_name'] = cell_col3.value
        elif (cell_col1.value == KEY.END_MAPPING):
            mapping_configs.append(config)
            config = {
                'collection_name' : None,
                'sheet_name' : None,
                'fields' : [],
            }
        else:
            field = {
                'column': cell_col1.value,
                'field_name' : cell_col2.value,
                'is_key': True if cell_col3.value == KEY.KEY_FIELD else False
            }
            config['fields'].append(field)
    return mapping_configs